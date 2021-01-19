from flask import make_response
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.writer.excel import save_virtual_workbook

from app.models.base_line import Patient, IniDiaPro, PastHis
from app.models.crf_info import FollInfo
from app.models.cycle import MoleDetec, SideEffect
from app.models.therapy_record import Surgery, TreRec, DetailTrePlan, Radiotherapy, OneToFive
from app.utils.date import get_age_by_birth, get_birth_date_by_id_card


class Export:
    pids = []
    header = []

    # 表头
    base_info_header = ['编号','身份证号' ,'姓名' , '性别' ,'出生日期','年龄','电话号码','PS评分']
    diagonse_header = ['吸烟史','饮酒史','初诊日期', '活检方式', '标本部位', '病理报告日期', '病理号',
        '病理诊断','c分期T','c分期N','c分期M', '临床分期','p分期T','p分期N','p分期M','病理分期', '转移部位']
    # gene_header = ['阳性基因','PD-L1表达水平','TMB']
    # immune_header = []
    surger_therapy_header = ['手术时间','手术范围','术后病理','病理号','病理诊断','术后辅助化疗','辅助治疗开始时间','辅助治疗结束时间','副反应','阳性基因','PD-L1表达水平','TMB']
    radio_therapy_header = ['放疗部位','放疗剂量',	'分割次数','放疗开始时间','放疗结束时间','疗效评价','副反应','阳性基因','PD-L1表达水平','TMB']
    survival_header = ['生存状态','死亡时间','最后一次随访日期']
    nth_therapy_header = ['治疗方案','开始日期','结束日期','疗效评估','副反应','进展日期','进展描述','阳性基因','PD-L1表达水平','TMB']

    # 缓存数据
    buffer = {}

    # map映射关系
    gender_map = {0: "女", 1: "男", "/": "/"}                   # 性别map
    radio_dos_unit_map = {0: "Gy", 1: "cGy"}                   # 放射剂量单位map
    survival_status_map = {1 : "死亡", 2 : "存活", 3 : "失联"}   # 生存状态map
    patDia_map = {
            "0-0": "上皮型肿瘤",
            "0-0-0": "上皮型肿瘤-腺癌",
            "0-0-0-0": "上皮型肿瘤-腺癌-贴壁型腺癌",
            "0-0-0-1": "上皮型肿瘤-腺癌-腺泡型腺癌",
            "0-0-0-2": "上皮型肿瘤-腺癌-乳头型腺癌",
            "0-0-0-3": "上皮型肿瘤-腺癌-微乳头型腺癌",
            "0-0-0-4": "上皮型肿瘤-腺癌-实体型腺癌",
            "0-0-0-5": "上皮型肿瘤-腺癌-浸润性粘液腺癌",
            "0-0-0-5-0": "上皮型肿瘤-腺癌-浸润性粘液腺癌-浸润和非浸润性混合型粘液性腺癌",
            "0-0-0-6": "上皮型肿瘤-腺癌-胶质性腺癌",
            "0-0-0-7": "上皮型肿瘤-腺癌-胎儿型腺癌",
            "0-0-0-8": "上皮型肿瘤-腺癌-肠型腺癌",
            "0-0-0-9": "上皮型肿瘤-腺癌-微小浸润性腺癌",
            "0-0-0-9-0": "上皮型肿瘤-腺癌-微小浸润性腺癌-非粘液性腺癌",
            "0-0-0-9-1": "上皮型肿瘤-腺癌-微小浸润性腺癌-粘液癌",
            "0-0-0-10": "上皮型肿瘤-腺癌-侵袭前病变",
            "0-0-0-10-0": "上皮型肿瘤-腺癌-侵袭前病变-非典型腺瘤样增生",
            "0-0-0-10-1": "上皮型肿瘤-腺癌-侵袭前病变-原位腺癌",
            "0-0-0-10-1-0": "上皮型肿瘤-腺癌-侵袭前病变-原位腺癌-非粘液性原位腺癌",
            "0-0-0-10-1-1": "上皮型肿瘤-腺癌-侵袭前病变-原位腺癌-粘液性原位腺癌",
            "0-0-1": "上皮型肿瘤-鳞癌",
            "0-0-1-0": "上皮型肿瘤-鳞癌-角化型鳞癌",
            "0-0-1-1": "上皮型肿瘤-鳞癌-非角化型鳞癌",
            "0-0-1-2": "上皮型肿瘤-鳞癌-基底鳞状细胞癌",
            "0-0-1-3": "上皮型肿瘤-鳞癌-侵袭前病变",
            "0-0-1-3-0": "上皮型肿瘤-鳞癌-侵袭前病变-鳞状细胞原位癌",
            "0-1": "神经内分泌肿瘤",
            "0-1-0": "神经内分泌肿瘤-小细胞癌",
            "0-1-0-0": "神经内分泌肿瘤-小细胞癌-结合小细胞癌",
            "0-1-1": "神经内分泌肿瘤-大细胞神经内分泌癌",
            "0-1-1-0": "神经内分泌肿瘤-大细胞神经内分泌癌-结合大细胞神经内分泌癌",
            "0-1-2": "神经内分泌肿瘤-类癌肿瘤",
            "0-1-2-0": "神经内分泌肿瘤-类癌肿瘤-典型类癌肿瘤",
            "0-1-2-1": "神经内分泌肿瘤-类癌肿瘤-非典型类癌肿瘤",
            "0-1-3": "神经内分泌肿瘤-侵袭前的病变",
            "0-1-3-0": "神经内分泌肿瘤-侵袭前的病变-弥漫性特发性肺神经内分泌细胞增生",
            "0-1-4": "神经内分泌肿瘤-大细胞癌",
            "0-1-5": "神经内分泌肿瘤-腺鳞癌",
            "0-1-6": "神经内分泌肿瘤-癌肉瘤样癌",
            "0-1-6-0": "神经内分泌肿瘤-癌肉瘤样癌-多形性癌",
            "0-1-6-1": "神经内分泌肿瘤-癌肉瘤样癌-梭形细胞癌",
            "0-1-6-2": "神经内分泌肿瘤-癌肉瘤样癌-巨细胞癌",
            "0-1-6-3": "神经内分泌肿瘤-癌肉瘤样癌-癌肉瘤",
            "0-1-6-4": "神经内分泌肿瘤-癌肉瘤样癌-肺胚细胞瘤",
            "0-1-7": "神经内分泌肿瘤-其他未分类癌",
            "0-1-7-0": "神经内分泌肿瘤-其他未分类癌-淋巴上皮样癌",
            "0-1-7-1": "神经内分泌肿瘤-其他未分类癌-NUT肿瘤",
            "0-1-8": "神经内分泌肿瘤-唾液型肿瘤",
            "0-1-8-0": "神经内分泌肿瘤-唾液型肿瘤-粘液表皮样癌肿瘤",
            "0-1-8-1": "神经内分泌肿瘤-唾液型肿瘤-腺样囊性癌",
            "0-1-8-2": "神经内分泌肿瘤-唾液型肿瘤-上皮-肌上皮癌",
            "0-1-8-3": "神经内分泌肿瘤-唾液型肿瘤-多形性腺瘤",
            "0-1-9": "神经内分泌肿瘤-乳头状瘤",
            "0-1-9-0": "神经内分泌肿瘤-乳头状瘤-鳞状细胞乳头状癌",
            "0-1-9-0-0": "神经内分泌肿瘤-乳头状瘤-鳞状细胞乳头状癌-外生型",
            "0-1-9-0-1": "神经内分泌肿瘤-乳头状瘤-鳞状细胞乳头状癌-逆向生长",
            "0-1-9-1": "神经内分泌肿瘤-乳头状瘤-腺型状瘤",
            "0-1-9-2": "神经内分泌肿瘤-乳头状瘤-腺鳞混合型乳头状瘤",
            "0-1-10": "神经内分泌肿瘤-腺瘤",
            "0-1-10-0": "神经内分泌肿瘤-腺瘤-良性硬化性肺细胞瘤",
            "0-1-10-1": "神经内分泌肿瘤-腺瘤-泡腺腺瘤",
            "0-1-10-2": "神经内分泌肿瘤-腺瘤-乳头状腺瘤",
            "0-1-10-3": "神经内分泌肿瘤-腺瘤-粘液性囊腺瘤腺瘤",
            "0-1-10-4": "神经内分泌肿瘤-腺瘤-粘液腺腺瘤",
            "0-2": "间叶性肿瘤",
            "0-2-0": "间叶性肿瘤-肺错构瘤",
            "0-2-1": "间叶性肿瘤-软骨瘤",
            "0-2-2": "间叶性肿瘤-PEComatous肿瘤",
            "0-2-2-0": "间叶性肿瘤-PEComatous肿瘤-淋巴管平滑肌瘤病",
            "0-2-2-1": "间叶性肿瘤-PEComatous肿瘤-PEComa-良性",
            "0-2-2-1-0": "间叶性肿瘤-PEComatous肿瘤-PEComa-良性-透明细胞瘤",
            "0-2-2-2": "间叶性肿瘤-PEComatous肿瘤-PEComa-恶性",
            "0-2-3": "间叶性肿瘤-先天性支气管周肌纤维母细胞肿瘤",
            "0-2-4": "间叶性肿瘤-弥漫性肺淋巴管瘤病",
            "0-2-5": "间叶性肿瘤-炎症性肌纤维母细胞瘤",
            "0-2-6": "间叶性肿瘤-上皮样血管内皮瘤",
            "0-2-7": "间叶性肿瘤-胸膜肺母细胞瘤",
            "0-2-8": "间叶性肿瘤-滑膜肉瘤",
            "0-2-9": "间叶性肿瘤-肺动脉内膜肉瘤",
            "0-2-10": "间叶性肿瘤-肺黏液肉瘤伴EWSR1-CREB1易位",
            "0-2-11": "间叶性肿瘤-肌上皮肿瘤",
            "0-2-11-0": "间叶性肿瘤-肌上皮肿瘤-肌上皮瘤",
            "0-2-11-1": "间叶性肿瘤-肌上皮肿瘤-肌上皮癌",
            "0-2-12": "间叶性肿瘤-淋巴细胞组织细胞肿瘤",
            "0-2-13": "间叶性肿瘤-结外边缘区黏膜相关淋巴组织淋巴瘤（MALT淋巴瘤）",
            "0-2-14": "间叶性肿瘤-弥漫性大细胞淋巴瘤",
            "0-2-15": "间叶性肿瘤-淋巴瘤样肉芽肿",
            "0-2-16": "间叶性肿瘤-血管内大B细胞淋巴瘤",
            "0-2-17": "间叶性肿瘤-肺朗格罕细胞组织细胞增生症",
            "0-2-18": "间叶性肿瘤-Erdheim-Chester病",
            "0-3": "异位肿瘤",
            "0-3-0": "异位肿瘤-生殖细胞肿瘤",
            "0-3-0-0": "异位肿瘤-生殖细胞肿瘤-畸胎瘤-成熟",
            "0-3-0-1": "异位肿瘤-生殖细胞肿瘤-畸胎瘤-不成熟",
            "0-3-1": "异位肿瘤-肺内的胸腺瘤",
            "0-3-2": "异位肿瘤-黑色素瘤",
            "0-3-3": "异位肿瘤-脑膜瘤",
            "0-4": "转移性肿瘤",
            "0-5": "其他"
            }  #病理诊断map
    detail_therapy_map = {"Chemotherapy":"化疗","TargetedTherapy":"靶向治疗","ImmunityTherapy":"免疫治疗","AntivascularTherapy":"抗血管治疗"}
    beEffEva_map = {"1":"PD-进展","2":"SD-稳定","3":"PR-部分缓解","4":"CR-完全缓解","5":"术后未发现新病灶",
                    "PD-进展": "PD-进展","SD-稳定":"SD-稳定","PR-部分缓解":"PR-部分缓解","CR-完全缓解":"CR-完全缓解","术后未发现新病灶":"术后未发现新病灶"
                    } #因为这里数据库里存对数据很乱
    def __init__(self, pids):

        self.pids = pids
        self.__init_buffer()

        self.header_array = []
        self.header_array.append(self.base_info_header)
        self.header_array.append(self.diagonse_header)
        # self.header_array.append(self.gene_header)
        # self.header_array.append(self.immune_header)
        self.header_array.append(self.surger_therapy_header)
        self.header_array.append(self.radio_therapy_header)
        self.header_array.append(self.survival_header)

        self.base_info_chosen = []
        '''self.diagnose_chosen = []
        self.surger_therapy_chosen = []
        self.radio_therapy_chosen = []
        self.radio_therapy_chosen = []
        self.survival_chosen = []'''

        for item in self.header_array:
            self.header.extend(item)
        for i in range(1,6):
            self.header += [str(i) + "线" + item for item in self.nth_therapy_header]

    def work(self):
        wb = Workbook()
        ws = wb.active
        ws.title = '导出样本数据'
        # self.add_first_row(ws)
        ws.append(self.header)

        for pid in self.pids:
            ws.append(self.get_row_data(pid))

        wb.save('my_excel_demo1.xls')
        content = save_virtual_workbook(wb)
        resp = make_response(content)
        resp.headers["Content-Disposition"] = 'attachment; filename=samples.xlsx'
        resp.headers['Content-Type'] = 'application/x-xlsx'
        return resp

    def work_new(self, chosen_headers=[]):
        self.header = []
        self.base_info_header = []

        if 'patNumber' in chosen_headers:
            self.base_info_header.append('编号')
            self.base_info_chosen.append('patNumber')
        if 'idNumber' in chosen_headers:
            self.base_info_header.append('身份证号')
            self.base_info_chosen.append('idNumber')
        if 'patientName' in chosen_headers:
            self.base_info_header.append('姓名')
            self.base_info_chosen.append('patientName')
        if 'gender' in chosen_headers:
            self.base_info_header.append('性别')
            self.base_info_chosen.append('gender')
        if 'birthday' in chosen_headers:
            self.base_info_header.append('出生日期')
            self.base_info_chosen.append('birthday')
        if 'age' in chosen_headers:
            self.base_info_header.append('年龄')
            self.base_info_chosen.append('age')
        if 'phoneNumber' in chosen_headers:
            self.base_info_header.append('电话号码')
            self.base_info_chosen.append('phoneNumber')
        if 'PSScore' in chosen_headers:
            self.base_info_header.append('PS评分')
            self.base_info_chosen.append('PSScore')

        self.header.extend(self.base_info_header)
        if 'diagnose' in chosen_headers:
            self.header.extend(self.diagonse_header)
        if 'surger_therapy' in chosen_headers:
            self.header.extend(self.surger_therapy_header)
        if 'radio_therapy' in chosen_headers:
            self.header.extend(self.radio_therapy_header)
        if 'survival' in chosen_headers:
            self.header.extend(self.survival_header)
        if 'one_to_five' in chosen_headers:
            for i in range(1, 6):
                self.header += [str(i) + "线" + item for item in self.nth_therapy_header]

        wb = Workbook()
        ws = wb.active
        ws.title = '导出样本数据'
        ws.append(self.header)
        for pid in self.pids:
            data = []
            p = self.buffer.get('Patient').get(pid)
            init_pro = self.buffer.get('IniDiaPro').get(pid)
            for item in self.base_info_chosen:
                if item == 'gender':
                    data.append(self.gender_map.get(self.filter_none(p, item)))
                elif item == 'PSScore':
                    data.append(self.filter_none(init_pro, item))
                elif item == 'age':
                    age = get_age_by_birth(get_birth_date_by_id_card(getattr(p, 'idNumber')))
                    data.append(age if age else '/')
                elif item =='phoneNumber':
                    data.append(self.filter_none(p, 'phoneNumber1'))
                else:
                    data.append(self.filter_none(p, item))

            if 'diagnose' in chosen_headers:
                data.extend(self.get_diagnose_info(pid))
            if 'surger_therapy' in chosen_headers:
                data.extend(self.get_surgery(pid))
            if 'radio_therapy' in chosen_headers:
                data.extend(self.get_radio(pid))
            if 'survival' in chosen_headers:
                data.extend(self.get_survival_info(pid))
            if 'one_to_five' in chosen_headers:
                data.extend(self.get_one_To_Five(pid))

            ws.append(data)

        wb.save('test.xls')
        content = save_virtual_workbook(wb)
        resp = make_response(content)
        resp.headers["Content-Disposition"] = 'attachment; filename=samples.xlsx'
        resp.headers['Content-Type'] = 'application/x-xlsx'
        return resp

    def add_first_row(self,sheet):
        col_index = 1
        for item in self.header_array:
            sheet.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=col_index+len(item)-1)
            col_index = col_index + len(item)

        for i in range(1,6):
            sheet.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=col_index + len(self.nth_therapy_header) - 1)
            col_index = col_index + len(self.nth_therapy_header)

        sheet['A1'].value = '基本信息'
        sheet['I1'].value = '诊断'
        sheet['Z1'].value = '基因检测'
        sheet['AA1'].value = '免疫相关指标'
        sheet['AC1'].value = '手术治疗'
        sheet['AL1'].value = '放疗'
        sheet['AS1'].value = '生存'
        sheet['AV1'].value = '一线治疗'
        sheet['BC1'].value = '二线治疗'
        sheet['BJ1'].value = '三线治疗'
        sheet['BQ1'].value = '四线治疗'
        sheet['BX1'].value = '五线治疗'

        for cell in [sheet['A1'],sheet['I1'],sheet['Z1'],sheet['AA1'],sheet['AC1'],
                     sheet['AL1'],sheet['AS1'],sheet['AV1'],sheet['BC1'],sheet['BJ1'],sheet['BQ1'],sheet['BX1']]:
            cell.alignment = Alignment(horizontal='center', vertical='center')






    def get_row_data(self,pid):
        row_data = []
        row_data.extend(self.get_base_info(pid))
        row_data.extend(self.get_diagnose_info(pid))
        # row_data.append(self.get_gene_info(pid, 0))
        # row_data.extend(self.get_immune_index(pid, 0))
        row_data.extend(self.get_surgery(pid))
        row_data.extend(self.get_radio(pid))
        row_data.extend(self.get_survival_info(pid))
        row_data.extend(self.get_one_To_Five(pid))

        return row_data

    def __init_buffer(self):
        patient_array = Patient.query.filter(Patient.id.in_(self.pids),Patient.is_delete==0).all()
        iniDiaPro_array = IniDiaPro.query.filter(IniDiaPro.pid.in_(self.pids),IniDiaPro.is_delete==0).all()
        pastHis_array = PastHis.query.filter(PastHis.pid.in_(self.pids),PastHis.is_delete==0).all()
        gene_array = MoleDetec.query.filter(MoleDetec.pid.in_(self.pids),MoleDetec.is_delete==0).all()
        surgery_array = Surgery.query.filter(Surgery.pid.in_(self.pids),Surgery.is_delete==0).all()
        treRec_array = TreRec.query.filter(TreRec.pid.in_(self.pids),TreRec.is_delete==0).all()
        detail_trePlan_array = DetailTrePlan.query.filter(DetailTrePlan.pid.in_(self.pids),DetailTrePlan.is_delete==0).all()
        radio_array = Radiotherapy.query.filter(Radiotherapy.pid.in_(self.pids),Radiotherapy.is_delete==0).all()
        side_effect_array = SideEffect.query.filter(SideEffect.pid.in_(self.pids),SideEffect.is_delete==0).all()
        follinfo_array = FollInfo.query.filter(FollInfo.pid.in_(self.pids), FollInfo.is_delete == 0).all()
        self.add_buffer('Patient', self.classify_by_pid(patient_array))
        self.add_buffer('IniDiaPro', self.classify_by_pid(iniDiaPro_array))
        self.add_buffer('PastHis', self.classify_by_pid(pastHis_array))
        self.add_buffer('MoleDetec', self.classify_by_treNum(gene_array))
        self.add_buffer('Surgery', self.classify_by_treNum(surgery_array))
        self.add_buffer('TreRec', self.classify_by_treNum(treRec_array))
        self.add_buffer('DetailTrePlan', self.array_classify_by_treNum(detail_trePlan_array))
        self.add_buffer('Radiotherapy', self.classify_by_treNum(radio_array))
        self.add_buffer('SideEffect', self.array_classify_by_treNum(side_effect_array))
        self.add_buffer('FollInfo', self.array_classify_by_pid(follinfo_array))
        self.add_nth_therapy_buffer(treRec_array)

    #传入tre_recs，找出oneToFive的
    def add_nth_therapy_buffer(self,tre_recs):
        items = OneToFive.query.filter_by().all()
        one_to_five_dict = {}
        for item in items:
            pid = getattr(item,'pid')
            if one_to_five_dict.get(pid) is None:
                one_to_five_dict[pid] = {}
            one_to_five_dict[pid][item.treNum] = item

        data = {}
        for tre_rec in tre_recs:
            pid = tre_rec.pid
            if data.get(pid) is None:
                data[pid] = {}
            if tre_rec.trement in ["one","two","three","four","five"]:
                if one_to_five_dict.get(pid):
                    item = one_to_five_dict.get(pid).get(tre_rec.treNum)
                    if item:
                        data[pid][tre_rec.trement] = item
        self.add_buffer('OneToFive',data)

    def add_detail_trePlan_buffer(self):
        pass

    # 缓存数据到内存中, 方便查询
    def add_buffer(self,key,value):
        self.buffer[key] = value

    def classify_by_pid(self,items):
        data = {}
        for item in items:
            if getattr(item,'pid',None):
                data[item.pid] = item
            else:
                data[item.id] = item
        return data

    def array_classify_by_pid(self,items):
        data = {}
        for item in items:
            pid = getattr(item,'pid')
            if data.get(pid) is None:
                data[pid] = []
            data[pid].append(item)
        return data

    # pid有多个，需要按照treNum分类。
    def classify_by_treNum(self,items):
        '''
        :param items:[objent]
        :return: map = {
            "pid":{
                "total":xxx,
                "treNum":{}
            }
        }
        '''
        data = {}
        for item in items:
            pid = getattr(item,'pid')
            if data.get(pid) is None:
                data[pid] = {}
                data[pid]['total'] = 0
            data[pid][item.treNum] = item
            data[pid]['total'] += 1
        return data

    def array_classify_by_treNum(self,items):
        '''
        :param items:[objent]
        :return: map = {
            "pid":{
                "total":xxx,
                "treNum":[]
            }
        }
        '''
        data = {}
        for item in items:
            pid = getattr(item,'pid')
            if data.get(pid) is None:
                data[pid] = {}
                data[pid]['total'] = 0
            if data.get(pid).get(item.treNum) is None:
                data[pid][item.treNum] = []
                data[pid]['total'] += 1
            data[pid][item.treNum].append(item)

        return data

    # 基本信息
    def get_base_info(self,pid):
        cnt = len(self.base_info_header)
        data = ['/'] * cnt

        p = self.buffer.get('Patient').get(pid)
        if p is None:
            print('p is none')
            return data
        init_pro = self.buffer.get('IniDiaPro').get(pid)
        data[0] = self.filter_none(p,'patNumber')
        data[1] = self.filter_none(p,'idNumber')
        data[2] = self.filter_none(p,'patientName')
        data[3] = self.gender_map.get(self.filter_none(p,'gender'))
        data[4] = self.filter_none(p,'birthday')
        age = get_age_by_birth(get_birth_date_by_id_card(getattr(p,'idNumber')))
        data[5] = age if age else '/'
        data[6] = self.filter_none(p,'phoneNumber1')
        data[7] = self.filter_none(init_pro,'PSScore') #PS评分
        if data[7] == -1: #远古时期的数据，-1代表空
            data[7] = '/'
        return data

    # 诊断
    def get_diagnose_info(self,pid):
        cnt = len(self.diagonse_header)
        data = ['/'] * cnt
        diagnose = self.buffer.get('PastHis').get(pid)
        init_pro = self.buffer.get('IniDiaPro').get(pid)

        data[0] = self.format_smoke_history(diagnose) #吸烟史
        data[1] = self.format_drink_history(diagnose) #饮酒史
        data[2] = self.filter_none(init_pro,'firVisDate')   #初诊日期
        data[3] = self.format_radio_data(init_pro,'bioMet') #活检方式
        data[4] = self.filter_none(init_pro,'speSite')      #标本部位
        data[5] = self.filter_none(init_pro,'patReDate')    #病理报告日期
        data[6] = self.filter_none(init_pro,'patNum')       #病理号
        data[7] = self.format_patDia(init_pro)    #病理诊断

        c_stage_array = self.format_stage(init_pro,'cStage')
        p_stage_array = self.format_stage(init_pro,'pStage')
        data[8] = c_stage_array[0]    #c分期T
        data[9] = c_stage_array[1]    #c分期N
        data[10] = c_stage_array[2]   #c分期M
        data[11] = self.filter_none(init_pro,'cliStage')  #c临床分期
        data[12] = p_stage_array[0]   #p分期T
        data[13] = p_stage_array[1]   #p分期N
        data[14] = p_stage_array[2]   #p分期M
        data[15] = self.filter_none(init_pro,'patStage')  #p病理分期
        data[16] = self.format_radio_data(init_pro,'traSite') #转移部位

        return data

    # 基因检测
    def get_gene_info(self,pid,treNum):
        mole_detecs = self.buffer.get('MoleDetec').get(pid)
        if mole_detecs is None:
            return '/'
        mole_detec = mole_detecs.get(treNum)
        if mole_detec is None:
            return '/'
        gene_record = self.get_positive_gene(mole_detec)
        if gene_record == []:
            return '/'
        value = ','.join(gene_record)
        return value

    # 免疫相关指标 pdl1 和 TMB
    def get_immune_index(self,pid,treNum):
        mole_detecs = self.buffer.get('MoleDetec').get(pid)
        data = ['/'] * 2
        if mole_detecs is None:
            return data
        mole_detec = mole_detecs.get(treNum)
        pdl1_val = '/'
        tmb_val = '/'
        if mole_detec and mole_detec.PDL1:
            pdl1_val = str(mole_detec.PDL1) + '%'
            tmb_val = self.filter_none(mole_detec.TMB)
            if mole_detec.PDL1KT:
                pdl1_val += ',抗体：' + str(mole_detec.PDL1KT)

        data[0] = pdl1_val
        data[1] = tmb_val
        return data

    def get_surgery(self,pid):
        tre_recs = self.buffer.get('TreRec').get(pid)
        treNum = None
        if tre_recs is None:
            return self._get_surgery_info(pid,treNum)
        for i in range(1,tre_recs.get('total')+1):
            tre_rec = tre_recs.get(i)
            if tre_rec and tre_rec.trement == 'surgery':
                treNum = tre_rec.treNum
                break
        return self._get_surgery_info(pid,treNum)

    # 获取treNum的手术
    def _get_surgery_info(self,pid,treNum):
        surgery_dict = self.buffer.get('Surgery').get(pid)
        tre_plan_dict = self.buffer.get('DetailTrePlan').get(pid)
        data = ['/'] * len(self.surger_therapy_header)
        if surgery_dict is None or tre_plan_dict is None:
            return data
        surgery = surgery_dict.get(treNum)
        tre_plans = tre_plan_dict.get(treNum)
        if surgery is None or surgery.posAdjChem == False:
            return data
        if tre_plans:
            tre_plans = sorted(tre_plans, key=lambda item: item.currPeriod if item.currPeriod is not None else -1) #TODO bug?
        else:
            tre_plans = []
        data[0] = self.filter_none(surgery.surDate)                                             #手术时间
        data[1] = self.format_radio_data(surgery,'surSco')                                      #手术范围
        data[2] = self.filter_none(self.change_bool_to_yes_or_no(surgery.isRepBio))             #术后病理
        data[3] = self.filter_none(self.change_bool_to_yes_or_no(surgery.specNum))              #病理号
        data[4] = self.format_patDia(surgery)                                                   #病理诊断
        data[5] = self.filter_none(self.change_bool_to_yes_or_no(tre_plans != []))              #术后辅助化疗
        data[6] = self.filter_none(tre_plans[0],'begDate')  if tre_plans != [] else '/'         #辅助治疗开始时间
        data[7] = self.filter_none(tre_plans[-1],'endDate') if tre_plans != [] else '/'         #辅助治疗结束时间
        data[8] = self.get_side_effect(pid,treNum)                                              #副反应
        data[9] = self.get_gene_info(pid, treNum)                                               #阳性基因

        immune_array = self.get_immune_index(pid, treNum)
        data[10] = immune_array[0]
        data[11] = immune_array[1]

        return data

    def get_radio(self,pid):
        tre_recs = self.buffer.get('TreRec').get(pid)
        treNum = None
        if tre_recs is None:
            return self._get_radio_info(pid,treNum)
        for i in range(1,tre_recs.get('total')+1):
            tre_rec = tre_recs.get(i)
            if tre_rec and tre_rec.trement == 'radiotherapy':
                treNum = tre_rec.treNum
                break
        return self._get_radio_info(pid,treNum)

    # 获取treNum的放疗
    def _get_radio_info(self,pid,treNum):
        radio_dict = self.buffer.get('Radiotherapy').get(pid)
        data = ['/'] * len(self.radio_therapy_header)
        if radio_dict is None :
            return data
        radio = radio_dict.get(treNum)
        if radio is None:
            return data
        #
        data[0] = self.format_radio_data(radio,'radSite')             #放疗部位
        data[1] = self.get_radio_dose(radio.radDose,radio.dosUnit)    #放疗剂量
        data[2] = self.value_with_unit(radio.splTim,radio.method)     #分割次数
        data[3] = self.filter_none(radio.begDate)                     #放疗开始时间
        data[4] = self.filter_none(radio.endDate)                     #放疗结束时间
        data[5] = self.get_effect_evaluation(pid,treNum)              #疗效评价
        data[6] = self.get_side_effect(pid,treNum)                    #副反应
        data[7] = self.get_gene_info(pid, treNum)                     #阳性基因

        immune_array = self.get_immune_index(pid, treNum)
        data[8] = immune_array[0]                                     #pdl1
        data[9] = immune_array[1]                                     #tmb

        return data

    def get_radio_dose(self,value,unit):
        if value is None:
            return '/'
        unit = self.radio_dos_unit_map.get(unit)
        return self.value_with_unit(value,unit)

    def value_with_unit(self,value,unit):
        if value is None:
            return None
        data = str(value)
        if unit:
            data += " " + str(unit)
        return data

    def get_patDia_value(self,json_value):
        if json_value is None or json_value == {}:
            return json_value

    # 生存 多条数据，导出最后的
    def get_survival_info(self,pid):
        data = ['/'] * len(self.survival_header)
        follInfo_array = self.buffer.get('FollInfo').get(pid)
        if follInfo_array is None:
            return data
        follInfo = follInfo_array[-1]
        data[0] = self.filter_none(self.survival_status_map.get(follInfo.livSta)) #生存状态
        data[1] = self.filter_none(follInfo.dieDate) #死亡日期
        data[2] = self.filter_none(follInfo.date) #随访日期
        return data

    def format_smoke_history(self,object):
        if object is None:
            return '/'
        is_smoke = object.smoke
        info = object.smokingHis
        value = '/'
        if is_smoke == 1:
            value = '累积吸烟时间（年）' + str(info.get('smokeYearAvg')) + '日平均吸烟量（支）' + str(info.get('smokeDayAvg'))
        elif is_smoke == 0:
            value = '无'
        return value

    def format_drink_history(self,object):
        if object is None:
            return '/'
        is_drink = object.drink
        info = object.drinkingHis
        value = '/'
        if is_drink == 1:
            value = '累积饮酒时间（年）' + str(info.get('drinkYearAvg')) + '日平均饮酒量（mL）' + str(info.get('drinkDayAvg'))
        elif is_drink == 0:
            value = '无'
        return value

    def format_radio_data(self,object,item):
        if object is None:
            return '/'
        data = getattr(object,item)
        if data is None:
            return '/'
        radio = data.get('radio')
        other = data.get('other')
        value = ",".join(radio)
        if other:
            value = value + "," + "其他: " + other
        return value

    def change_bool_to_yes_or_no(self,bool_value):
        if bool_value is None:
            return None
        return '是' if bool_value else '否'

    # 既能过滤空数据，也能过滤对象是空对象的情况
    def filter_none(self,data,item=None):
        if data is None:
            return '/'
        if item:
            val = getattr(data,item)
        else:
            val = data
        return val if val is not None else '/'

    # 格式化C分期,P分期
    def format_stage(self,object,item):
        value = ['/'] * 3
        if object is None:
            return value
        data = getattr(object,item)
        if data is None or data == "":
            return value
        ret = data.split(',')
        for i in range(0,3):
            value[i] = self.filter_none(ret[i])
        return value

    # 格式化病理诊断
    def format_patDia(self,object):
        if object is None or object.patDia is None:
            return '/'
        patDia = object.patDia
        radio = patDia.get('radio')
        data = []
        for item in radio:
            if item != "0-5":
                data.append(self.patDia_map.get(item))
            else:
                data.append("其他：" + str(patDia.get('other')))
        return ','.join(data)

    # 获取某一分子检测的阳性基因
    def get_positive_gene(self,mole_detec):
        if mole_detec is None:
            return []
        gene_record = []
        if mole_detec.ALK == 1:
            gene_record.append('ALK')
        if mole_detec.BIM == 1:
            gene_record.append('BIM')
        if mole_detec.BRAF == 1:
            gene_record.append('BRAF')
        if mole_detec.cMET == 1:
            gene_record.append('cMET')
        if mole_detec.EGFR == 1:
            gene_record.append('EGFR')
        if mole_detec.HER_2 == 1:
            gene_record.append('HER_2')
        if mole_detec.KRAS == 1:
            gene_record.append('KRAS')
        if mole_detec.PIK3CA == 1:
            gene_record.append('PIK3CA')
        if mole_detec.ROS1 == 1:
            gene_record.append('ROS1')
        if mole_detec.RET == 1:
            gene_record.append('RET')
        if mole_detec.UGT1A1 == 1:
            gene_record.append('UGT1A1')
        return gene_record

    # 获取副反应
    def get_side_effect(self,pid,treNum):
        side_effect_dict = self.buffer.get('SideEffect').get(pid)
        if side_effect_dict is None:
            return '/'
        side_effects =  side_effect_dict.get(treNum)
        if side_effects is None:
            return '/'
        side_effects = sorted(side_effects,key= lambda item:item.id)
        data = []
        for side_effect in side_effects:
            if side_effect.sidReaName:
                data.append(side_effect.sidReaName)
        return ','.join(data)

    # 获取疗效评估
    def get_effect_evaluation(self,pid,treNum):
        tre_rec_dict = self.buffer.get('TreRec').get(pid)
        if tre_rec_dict is None:
            return '/'
        tre_rec = tre_rec_dict.get(treNum)
        if tre_rec is None:
            return '/'
        value = self.beEffEva_map.get(tre_rec.beEffEva)
        return value if value else '/'

    # 获取一线到五线到信息
    def get_one_To_Five(self,pid):
        one_to_five = ['one','two',"three","four","five"]
        data = []
        for nth in one_to_five:
            data.extend(self._get_nth_therapy(pid,nth))
        return data

    # 获取第几线治疗
    #['治疗方案','开始日期','结束日期','疗效评估','副反应','进展日期','进展描述']
    def _get_nth_therapy(self,pid,nth):
        data = ['/'] * len(self.nth_therapy_header)
        one_to_five_dict = self.buffer.get('OneToFive').get(pid)
        treRec_dict = self.buffer.get('TreRec').get(pid)

        if one_to_five_dict is None:
            return data
        therapy = one_to_five_dict.get(nth)
        if therapy is None:
            return data
        treRec = None
        treNum = therapy.treNum
        if treRec_dict:
            treRec = treRec_dict.get(treNum)
        data[0] = self.get_therapy_plan(pid,treNum)     #治疗方案
        data[1] = self.filter_none(therapy.begDate)     #开始日期
        data[2] = self.filter_none(therapy.endDate)      #结束日期
        data[3] = self.get_effect_evaluation(pid,treNum) #疗效评估
        data[4] = self.get_side_effect(pid,treNum)       #副反应
        data[5] = self.filter_none(treRec,"proDate")    #进展日期
        data[6] =  self.filter_none(treRec,"proDes")    #进展描述
        data[7] = self.get_gene_info(pid, treNum)  # 阳性基因

        immune_array = self.get_immune_index(pid, treNum)
        data[8] = immune_array[0]       #pdl1
        data[9] = immune_array[1]       #tmb

        return data

    #获取1到5线到治疗方案
    def get_therapy_plan(self,pid,treNum):
        treRec_dict = self.buffer.get('TreRec').get(pid)
        if treRec_dict is None:
            return '/'
        tre_rec = treRec_dict.get(treNum)
        if tre_rec is None:
            return '/'
        if tre_rec.trement is None:
            return '/'
        therapy_plans = tre_rec.trement.split(',')
        data = []
        for plan in therapy_plans:
            temp = self.detail_therapy_map.get(plan)
            if temp:
                data.append(temp)
        if data == []:
            return '/'
        return ",".join(data)
