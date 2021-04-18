from time import time

from flask import make_response
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from app.models.base_line import DrugHistory, Patient, IniDiaPro, PastHis, SpecimenInfo
from app.models.crf_info import FollInfo
from app.models.cycle import Signs, SideEffect, MoleDetec, Immunohis
from app.models.lab_inspectation import BloodRoutine, BloodBio, Thyroid, Coagulation, MyocardialEnzyme, Cytokines, \
    LymSubsets, UrineRoutine, TumorMarker
from app.models.other_inspect import Lung, OtherExams, ImageExams
from app.models.therapy_record import DetailTrePlan, Surgery, OneToFive, Radiotherapy, TreRec
from app.spider.research_center import ResearchCenterSpider
import gc


class Export:
    # 缓存数据
    buffer = {}

    def __init__(self, pids, treNums, follInfoNum, baseLine, trementInfo, follInfo):
        self.pids = pids
        self.treNums = treNums
        self.follInfoNum = follInfoNum
        self.baseLine = baseLine
        self.trementInfo = trementInfo
        self.follInfo = follInfo

        self.headers = []
        if follInfo:
            self.if_follInfo = True
        else:
            self.if_follInfo = False

        self.__init_buffer()

    def work(self):
        for treNum in self.treNums:
            if treNum == 0:
                for dic in self.baseLine:
                    obj_class_name = dic['table']
                    columns = dic['column']
                    obj = obj_class_name()
                    self.headers.extend(obj.get_export_header(columns, self.buffer))
            else:
                if self.trementInfo:
                    self.headers.append('治疗信息' + str(treNum))
                for dic in self.trementInfo:
                    obj_class_name = dic['table']
                    columns = dic['column']
                    obj = obj_class_name()
                    self.headers.extend(obj.get_export_header(columns, self.buffer))

        if self.follInfo:
            obj = FollInfo()
            self.headers.extend(obj.get_export_header(self.follInfo['column'], self.buffer, self.follInfoNum))

        # wb = Workbook()
        wb = Workbook(write_only=True)
        # ws = wb.active
        ws = wb.create_sheet(title='导出样本数据')
        # ws.title = '导出样本数据'
        ws.append(self.headers)

        for pid in self.pids:
            row = []
            for treNum in self.treNums:
                if treNum == 0:
                    for dic in self.baseLine:
                        obj_class_name = dic['table']
                        columns = dic['column']
                        obj = obj_class_name()
                        row.extend(obj.get_export_row(columns, self.buffer, pid, treNum))
                else:
                    if self.trementInfo:
                        row.append('')
                    for dic in self.trementInfo:
                        obj_class_name = dic['table']
                        columns = dic['column']
                        obj = obj_class_name()
                        row.extend(obj.get_export_row(columns, self.buffer, pid, treNum))
            if self.follInfo:
                obj = FollInfo()
                row.extend(obj.get_export_row(self.follInfo['column'], self.buffer, pid, 0, self.follInfoNum))
            ws.append(row)
            gc.collect()

        content = save_virtual_workbook(wb)
        resp = make_response(content)
        resp.headers["Content-Disposition"] = 'attachment; filename=samples.xlsx'
        resp.headers['Content-Type'] = 'application/x-xlsx'
        return resp

    def __init_buffer(self):
        # 缓存研究中心信息
        centers_info = ResearchCenterSpider().search_all().get('data')
        centers_dict = {}
        for center in centers_info:
            id = center.get('id')
            centers_dict[id] = center.get('name')
        self.add_buffer('ResearchCenter', centers_dict)

        patient_array = Patient.query.filter(Patient.id.in_(self.pids), Patient.is_delete == 0).all()
        iniDiaPro_array = IniDiaPro.query.filter(IniDiaPro.pid.in_(self.pids), IniDiaPro.is_delete == 0).all()
        pastHis_array = PastHis.query.filter(PastHis.pid.in_(self.pids), PastHis.is_delete == 0).all()
        drugHistory_array = DrugHistory.query.filter(DrugHistory.pid.in_(self.pids), DrugHistory.is_delete == 0).all()
        specimenInfo_array = SpecimenInfo.query.filter(SpecimenInfo.pid.in_(self.pids), SpecimenInfo.is_delete == 0).all()

        treRec_array = TreRec.query.filter(TreRec.pid.in_(self.pids), TreRec.treNum.in_(self.treNums),
                                           TreRec.is_delete == 0).all()
        surgery_array = Surgery.query.filter(Surgery.pid.in_(self.pids), Surgery.treNum.in_(self.treNums),
                                             Surgery.is_delete == 0).all()
        oneToFive_array = OneToFive.query.filter(OneToFive.pid.in_(self.pids), OneToFive.treNum.in_(self.treNums),
                                                 OneToFive.is_delete == 0).all()
        radio_array = Radiotherapy.query.filter(Radiotherapy.pid.in_(self.pids), Radiotherapy.treNum.in_(self.treNums),
                                                Radiotherapy.is_delete == 0).all()
        detail_trePlan_array = DetailTrePlan.query.filter(DetailTrePlan.pid.in_(self.pids),
                                                          DetailTrePlan.treNum.in_(self.treNums),
                                                          DetailTrePlan.is_delete == 0).all()
        signs_array = Signs.query.filter(Signs.pid.in_(self.pids), Signs.treNum.in_(self.treNums),
                                         Signs.is_delete == 0).all()
        sideEffect_array = SideEffect.query.filter(SideEffect.pid.in_(self.pids), SideEffect.treNum.in_(self.treNums),
                                                   SideEffect.is_delete == 0).all()

        bloodRoutine_array = BloodRoutine.query.filter(BloodRoutine.pid.in_(self.pids),
                                                       BloodRoutine.treNum.in_(self.treNums),
                                                       BloodRoutine.is_delete == 0).all()
        bloodBio_array = BloodBio.query.filter(BloodBio.pid.in_(self.pids), BloodBio.treNum.in_(self.treNums),
                                               BloodBio.is_delete == 0).all()
        thyroid_array = Thyroid.query.filter(Thyroid.pid.in_(self.pids), Thyroid.treNum.in_(self.treNums),
                                             Thyroid.is_delete == 0).all()
        coagulation_array = Coagulation.query.filter(Coagulation.pid.in_(self.pids),
                                                     Coagulation.treNum.in_(self.treNums),
                                                     Coagulation.is_delete == 0).all()
        myocardialEnzyme_array = MyocardialEnzyme.query.filter(MyocardialEnzyme.pid.in_(self.pids),
                                                               MyocardialEnzyme.treNum.in_(self.treNums),
                                                               MyocardialEnzyme.is_delete == 0).all()
        cytokines_array = Cytokines.query.filter(Cytokines.pid.in_(self.pids), Cytokines.treNum.in_(self.treNums),
                                                 Cytokines.is_delete == 0).all()
        lymSubsets_array = LymSubsets.query.filter(LymSubsets.pid.in_(self.pids), LymSubsets.treNum.in_(self.treNums),
                                                   LymSubsets.is_delete == 0).all()
        urineRoutine_array = UrineRoutine.query.filter(UrineRoutine.pid.in_(self.pids),
                                                       UrineRoutine.treNum.in_(self.treNums),
                                                       UrineRoutine.is_delete == 0).all()
        tumorMarker_array = TumorMarker.query.filter(TumorMarker.pid.in_(self.pids),
                                                     TumorMarker.treNum.in_(self.treNums),
                                                     TumorMarker.is_delete == 0).all()
        lung_array = Lung.query.filter(Lung.pid.in_(self.pids), Lung.treNum.in_(self.treNums),
                                       Lung.is_delete == 0).all()
        otherExams_array = OtherExams.query.filter(OtherExams.pid.in_(self.pids), OtherExams.treNum.in_(self.treNums),
                                                   OtherExams.is_delete == 0).all()
        imageExams_array = ImageExams.query.filter(ImageExams.pid.in_(self.pids), ImageExams.treNum.in_(self.treNums),
                                                   ImageExams.is_delete == 0).all()
        immunohis_array = Immunohis.query.filter(Immunohis.pid.in_(self.pids), Immunohis.treNum.in_(self.treNums),
                                                 Immunohis.is_delete == 0).all()
        moleDetec_array = MoleDetec.query.filter(MoleDetec.pid.in_(self.pids), MoleDetec.treNum.in_(self.treNums),
                                                 MoleDetec.is_delete == 0).all()
        follinfo_array = FollInfo.query.filter(FollInfo.pid.in_(self.pids), FollInfo.is_delete == 0).order_by(FollInfo.date.desc()).all()

        self.add_buffer('Patient', self.classify_by_pid(patient_array))
        self.add_buffer('PastHis', self.classify_by_pid(pastHis_array))
        self.add_buffer('DrugHistory', self.array_classify_by_pid(drugHistory_array))
        self.add_buffer('IniDiaPro', self.classify_by_pid(iniDiaPro_array))
        self.add_buffer('SpecimenInfo', self.array_classify_by_pid(specimenInfo_array))

        self.add_buffer('TreRec', self.classify_by_treNum(treRec_array))
        self.add_buffer('Surgery', self.classify_by_treNum(surgery_array))
        self.add_buffer('OneToFive', self.classify_by_treNum(oneToFive_array))
        self.add_buffer('DetailTrePlan', self.array_classify_by_treSolu(detail_trePlan_array))
        self.add_buffer('Radiotherapy', self.classify_by_treNum(radio_array))
        self.add_buffer('Signs', self.array_classify_by_treNum(signs_array))
        self.add_buffer('SideEffect', self.array_classify_by_treNum(sideEffect_array))

        self.add_buffer('BloodRoutine', self.classify_by_treNum(bloodRoutine_array))
        self.add_buffer('BloodBio', self.classify_by_treNum(bloodBio_array))
        self.add_buffer('Thyroid', self.classify_by_treNum(thyroid_array))
        self.add_buffer('Coagulation', self.classify_by_treNum(coagulation_array))
        self.add_buffer('MyocardialEnzyme', self.classify_by_treNum(myocardialEnzyme_array))
        self.add_buffer('Cytokines', self.classify_by_treNum(cytokines_array))
        self.add_buffer('LymSubsets', self.classify_by_treNum(lymSubsets_array))
        self.add_buffer('UrineRoutine', self.classify_by_treNum(urineRoutine_array))
        self.add_buffer('TumorMarker', self.classify_by_treNum(tumorMarker_array))
        self.add_buffer('Lung', self.classify_by_treNum(lung_array))
        self.add_buffer('OtherExams', self.classify_by_treNum(otherExams_array))
        self.add_buffer('ImageExams', self.array_classify_by_treNum(imageExams_array))
        self.add_buffer('Immunohis', self.classify_by_treNum(immunohis_array))
        self.add_buffer('MoleDetec', self.classify_by_treNum(moleDetec_array))

        self.add_buffer('FollInfo', self.array_classify_by_pid(follinfo_array))

    # 缓存数据到内存中, 方便查询
    def add_buffer(self, key, value):
        self.buffer[key] = value

    def classify_by_pid(self, items):
        data = {}
        for item in items:
            if getattr(item, 'pid', None):
                data[item.pid] = item
            else:
                data[item.id] = item
        return data

    def array_classify_by_pid(self, items):
        data = {}
        for item in items:
            pid = getattr(item, 'pid')
            if data.get(pid) is None:
                data[pid] = []
            data[pid].append(item)
        return data

    # pid有多个，需要按照treNum分类。
    def classify_by_treNum(self, items):
        '''
        :param items:[objent]
        :return: map = {
            pid:{
                treNum:obj
            }
        }
        '''
        data = {}
        for item in items:
            pid = getattr(item, 'pid')
            if data.get(pid) is None:
                data[pid] = {}
            if data.get(pid).get(item.treNum) is None:
                data[pid][item.treNum] = item
        return data

    def array_classify_by_treNum(self, items):
        '''
        :param items:[objent]
        :return: map = {
            pid:{
                treNum:[]
            }
        }
        '''
        data = {}
        for item in items:
            pid = getattr(item, 'pid')
            if data.get(pid) is None:
                data[pid] = {}
            if data.get(pid).get(item.treNum) is None:
                data[pid][item.treNum] = []
            data[pid][item.treNum].append(item)

        return data

    def array_classify_by_treSolu(self, items):
        '''
        :param items:[objent]
        :return: map = {
            pid:{
                treNum:{
                    treSolu:[..,..]
                }
            }
        }
        '''
        data = {}
        for item in items:
            pid = getattr(item, 'pid')
            if data.get(pid) is None:
                data[pid] = {}
            if data.get(pid).get(item.treNum) is None:
                data[pid][item.treNum] = {}
            if data.get(pid).get(item.treNum).get(item.treSolu) is None:
                data[pid][item.treNum][item.treSolu] = []
            data[pid][item.treNum][item.treSolu].append(item)

        return data